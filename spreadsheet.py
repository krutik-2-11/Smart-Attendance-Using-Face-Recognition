from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.cell import Cell
import time
import os
import sqlite3

#database connection
conn = sqlite3.connect('Student_DataBase.db')
c = conn.cursor()

#get current date
currentDate = time.strftime("%d_%m_%y")

#create a workbook and add a worksheet
if(os.path.exists('./Attendance.xlsx')):
    wb = load_workbook(filename = "Attendance.xlsx")
    sheet = wb.get_sheet_by_name('CseIII')
    #sheet = wb.active
    col_index = 1
    while(True):
        col = get_column_letter(col_index)
        col = col + '1'
        if(sheet[col].value == currentDate):
            break
        elif (sheet[col].value == None):
            sheet[col].value = currentDate
            break
        col_index += 1
        print('Ghusa')
    #entering info of only new students
    existing_roles = []
    col = get_column_letter(1)
    i = 3
    while(True):
        col = col + str(i)
        if(sheet[col].value == None):
            break
        existing_roles.append(sheet[col].value)
        col = get_column_letter(1)
        i += 1
    print(existing_roles)
    c.execute("SELECT enrollment,first_name,last_name FROM Attendance")
    
    while True:
        a = c.fetchone()
        print(a)
        if a == None:
            break
        
        else:
            if a[0] in existing_roles:
                print('Enrollment number already present')
            else:
                #sheet.append((a[0],a[1],a[2]))
                sheet['A' + str(i)] = a[0]
                sheet['B' + str(i)] = a[1]
                sheet['C' + str(i)] = a[2]
                existing_roles.append(a[0])
        

    wb.save(filename = "Attendance.xlsx")

   
        
else:
    wb = Workbook()
    dest_filename = 'Attendance.xlsx'
    c.execute("SELECT enrollment,first_name,last_name FROM Attendance ORDER BY enrollment")
    
    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = "CseIII"
    ws1.append(('Roll Number', 'First Name','Last Name', currentDate))
    ws1.append(('', '', '',''))

    #entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            ws1.append((a[0],a[1], a[2]))

    #saving the file
    wb.save(filename = dest_filename)

